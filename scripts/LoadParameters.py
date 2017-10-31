def ReadXl(xlfilename):
  from openpyxl import load_workbook
  from tqdm import tqdm
  import time

  wb = load_workbook(filename = xlfilename, read_only=True)
  Parms = {}

  for SheetName in wb.get_sheet_names():
    sheet = wb[SheetName]

    ItemList = []
    i = 2
    while sheet.cell(row=i, column=1).value:
      Item = {}
      j = 1
      while sheet.cell(row=1, column=j).value:
        Item[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        j += 1
      ItemList.append(Item)
      i += 1
    Parms[SheetName] = ItemList
  return Parms

if __name__ == '__main__':
  import json
  import sys

  argvs = sys.argv
  r = ReadXl(argvs[1])
  print json.dumps(r)
